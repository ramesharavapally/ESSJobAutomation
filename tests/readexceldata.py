import openpyxl
import json
import os
from datetime import datetime


def is_file_open(file_path):
    try:
        with open(file_path, 'a') as f:
            pass
        return False
    except PermissionError:
        return True    

def read_sheet_as_strings(file_path, sheet_name):
    # Load the workbook
    workbook = openpyxl.load_workbook(file_path, data_only=True)

    # Extract data from the specified sheet
    sheet = workbook[sheet_name]
    data_sheet = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = []
        for cell in row:
            if cell is None:
                row_data.append("")  # Treat None as an empty string
            else:
                if isinstance(cell , datetime):
                    cell = str(datetime.strftime(cell , '%m/%d/%y'))
                row_data.append(str(cell))  # Convert cell value to string
        data_sheet.append(row_data)

    return data_sheet

def lovs_excel_to_json(file_path):
        
    if is_file_open(file_path):
        raise Exception(f"Error: File '{file_path}' is already open. Please close the file and try again.")            
    sheet = read_sheet_as_strings(file_path= file_path , sheet_name='LOVS')
    
    # Define an empty list to store the data
    data = []
    
    # Iterate over rows in the Excel sheet, starting from the second row to skip the header    
    for row in sheet:
        # Create a dictionary to store each row of data
        row = list(row)
        for i in range(len(row)):
            if row[i] is None:
                row[i] = ""
        row_data = {
            'application_name': row[0],
            'user_lov_name': row[1],
            'description': row[2],
            'lov_type': row[3],
            'query': row[4]
        }
        # Append the row dictionary to the data list
        data.append(row_data)
    
    # Convert the data list
    return data

def lovs_standard_jobs_to_json(file_path):
        
    if is_file_open(file_path):
        raise Exception(f"Error: File '{file_path}' is already open. Please close the file and try again.")            
        
    sheet1 = read_sheet_as_strings(file_path= file_path , sheet_name='Header')
    # Define an empty list to store the data
    # data = []
    
    data_sheet1 = {}
    # for row in sheet1.iter_rows(min_row=2, values_only=True):
    for row in sheet1:
        row = list(row)
        for i in range(len(row)):
            if row[i] is None:
                row[i] = ""
        # print(row[0] , row[1])
        data_sheet1[row[0]] = row[1]
        
    # Extract data from Sheet2    
    sheet2 = read_sheet_as_strings(file_path=file_path , sheet_name='Parameters')
    parameters = []    
    for row in sheet2:
        row = list(row)
        for i in range(len(row)):
            if row[i] is None:
                row[i] = ""
        parameter = {
            "Parameter": row[0],
            "Display": row[1]            
        }
        parameters.append(parameter)
        
    # Construct the final JSON object    
    output_json = {
        **data_sheet1,
        "parameters": parameters
    }
    # print(output_json)
    
    return output_json

def jobs_excel_to_json(file_path):
    
    if is_file_open(file_path):
        raise Exception(f"Error: File '{file_path}' is already open. Please close the file and try again.")            

    # Extract data from Sheet1    
    sheet1 = read_sheet_as_strings(file_path=file_path , sheet_name='Header')
    print(sheet1)
    data_sheet1 = {}    
    for row in sheet1:
        row = list(row)
        for i in range(len(row)):
            if row[i] is None:
                row[i] = ""
        data_sheet1[row[0]] = row[1]

    # Extract data from Sheet2    
    sheet2 = read_sheet_as_strings(file_path=file_path , sheet_name='Parameters')
    parameters = []    
    for row in sheet2:
        row = list(row)
        for i in range(len(row)):
            if row[i] is None:
                row[i] = ""
        parameter = {
            "ParameterPrompt": row[0],
            "DataType": row[1],
            "NumberFormat": row[2],
            "ReadOnly": row[3],
            "value": row[4],
            "PageElement": row[5],
            "ListOfValueSource": row[6],
            "Attribute": row[7],
            "DefaultValue": row[8],
            "Required": row[9],
            "Tooltip": row[10],
            "DontDisplay": row[11],
            "ShowType":row[12],
            "DefaultDateFormat" :row[13],
            "DefaultDate" :row[14]
        }
        parameters.append(parameter)

    # Construct the final JSON object
    output_json = {
        **data_sheet1,
        "parameters": parameters
    }
    
    return output_json

def get_jobs_source_data(folder_path : str):                
    # List to store JSON data from all files
    json_data_array = []

    # Iterate over files in the folder
    for filename in os.listdir(folder_path):        
        if filename.endswith('.xlsx'):
            file_path = os.path.join(folder_path, filename)
            json_data = jobs_excel_to_json(file_path)
            json_data_array.append(json_data)

    # Convert the array to JSON format
    # json_data_array = [item.decode('utf-8') if isinstance(item, bytes) else item for item in json_data_array]
    # print(json.loads(json_data_array))
    output_json_str = json.dumps(json_data_array, indent=4)
    return output_json_str  

def get_lovs_source_data(folder_path : str):
    # List to store JSON data from all files
    json_data_array = []

    # Iterate over files in the folder
    for filename in os.listdir(folder_path):        
        if filename.endswith('.xlsx'):
            file_path = os.path.join(folder_path, filename)
            json_data = lovs_excel_to_json(file_path)
            json_data_array.append(json_data)

    # Convert the array to JSON format
    output = []
    for sublist in json_data_array:
        output.extend(sublist)
    output_json_str = json.dumps(output, indent=4)
    return output_json_str            

def get_standard_jobs_source_data(folder_path : str):
    # List to store JSON data from all files
    json_data_array = []        

    # Iterate over files in the folder
    for filename in os.listdir(folder_path):        
        if filename.endswith('.xlsx'):
            file_path = os.path.join(folder_path, filename)
            json_data = lovs_standard_jobs_to_json(file_path)
            json_data_array.append(json_data)

    # Convert the array to JSON format
    output = []
    # print(json_data_array)
    for sublist in json_data_array:        
        output.append(sublist)    
    output_json_str = json.dumps(output, indent=4)    
    return output_json_str            

# Example usage:
folder_path = '..\data\jobs'
json_data = get_jobs_source_data(folder_path)
print(json_data)

# folder_path = '..\data\lovs'
# json_data = get_lovs_source_data(folder_path)
# print(json_data)